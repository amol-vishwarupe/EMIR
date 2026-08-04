#!/usr/bin/env python3
"""
EMIR Refit trade validation engine.

Evaluates a trade CSV against reference_data/validation_rules.json, resolving
Referential-type rules against the mock GLEIF/DSB lookup files, and writes a
multi-sheet Excel results workbook.

Usage:
    python validate_and_report.py
    python validate_and_report.py --trades path/to/other_trades.csv --out path/to/results.xlsx

All paths default to the standard project layout (see README.md):
    <project>/trade_data_200.csv
    <project>/reference_data/validation_rules.json
    <project>/reference_data/gleif_lei_reference.json
    <project>/reference_data/dsb_upi_reference.json
    <project>/test_results.xlsx   (output)
"""
import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency 'openpyxl'. Install it with: python -m pip install openpyxl")

# =============================================================================
# Configuration that is NOT derivable from the data itself.
# In a real deployment these come from the reporting firm's own static config,
# not from the trade file or reference data.
SUBMITTING_ENTITY_LEI = "5493001KJTIIGC8Y1R12"  # R033: "our" LEI as report submitter

ACTIVE_CCY = {
    "USD", "EUR", "GBP", "JPY", "CHF", "AUD", "CAD", "NZD", "SEK", "NOK", "DKK",
    "SGD", "HKD", "CNY", "INR", "ZAR", "MXN", "BRL", "PLN", "CZK", "HUF", "TRY",
    "ILS", "KRW", "THB", "MYR", "IDR", "PHP", "RON", "BGN",
}
DEPRECATED_CCY = {
    "DEM", "FRF", "ITL", "ESP", "ATS", "BEF", "NLG", "PTE", "FIM", "GRD", "IEP",
    "LUF", "XEU", "SIT", "SKK", "CYP", "MTL", "EEK", "LVL", "LTL",
}

UTI_RE = re.compile(r"^[A-Z0-9]{18,52}$")
UPI_RE = re.compile(r"^[A-Z0-9]{12}$")
CCY_RE = re.compile(r"^[A-Z]{3}$")
ACTIONTYPE_ENUM = {"NEWT", "MODI", "CORR", "EROR", "REVI", "VALU", "POSC", "TERM", "ETRM", "COMP"}
ASSETCLASS_ENUM = {"CO", "CR", "CU", "EQ", "IR", "OT"}
UPI_MANDATORY_ACTIONS = {"NEWT", "MODI", "CORR", "REVI"}
PRIOR_NEWT_REQUIRED_ACTIONS = {"MODI", "CORR", "EROR", "REVI", "VALU", "POSC", "ETRM"}
LIFECYCLE_SUFFIX_RE = re.compile(
    r"-\d+-(NEWT|MODI|CORR|EROR|REVI|VALU|POSC|TERM|ETRM|COMP).*$"
)


# ---- helpers -----------------------------------------------------------------
def lei_check_digits(base18: str) -> str:
    s = base18 + "00"
    num = "".join(ch if ch.isdigit() else str(ord(ch.upper()) - ord("A") + 10) for ch in s)
    return f"{98 - (int(num) % 97):02d}"


def valid_lei_format(v: str) -> bool:
    if not v or not re.match(r"^[A-Z0-9]{18}[0-9]{2}$", v):
        return False
    return v[18:20] == lei_check_digits(v[:18])


def parse_date(v: str):
    try:
        return date.fromisoformat(v)
    except Exception:
        return None


def parse_dt(v: str):
    if not v or not re.match(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z$", v):
        return None
    try:
        return datetime.strptime(v, "%Y-%m-%dT%H:%M:%SZ")
    except Exception:
        return None


def is_number(v: str):
    try:
        return float(v)
    except Exception:
        return None


def lifecycle_key(trade_id: str) -> str:
    # Fixture-only heuristic for R021: strips a trailing "-<n>-<ACTIONTYPE>..."
    # lifecycle-step suffix so NEWT/MODI rows for "the same derivative" link up.
    # TradeID is not a regulatory field. Real systems need an explicit
    # PriorUTI / lifecycle-link field -- this does not generalize to arbitrary
    # production data. See README.md "Known limitations".
    return LIFECYCLE_SUFFIX_RE.sub("", trade_id)


# ---- rule engine ---------------------------------------------------------------
class Engine:
    def __init__(self, trades, gleif, dsb):
        self.trades = trades
        self.gleif = gleif
        self.dsb = dsb

        self.by_uti = defaultdict(list)
        for row in trades:
            self.by_uti[row["UTI"]].append(row)

        self.newt_by_generator = defaultdict(list)
        for row in trades:
            if row["ActionType"] == "NEWT" and row["UTI"]:
                self.newt_by_generator[(row["Counterparty1LEI"], row["UTI"])].append(row)

        self.by_lifecycle = defaultdict(list)
        for row in trades:
            self.by_lifecycle[lifecycle_key(row["TradeID"])].append(row)

        self.rules = {
            f"R{i:03d}": getattr(self, f"r{i:03d}") for i in range(1, 51)
        }

    # Each rule returns True (PASS), False (FAIL/WARN per rule severity),
    # or None (N/A -- not evaluable for this row, e.g. condition not applicable
    # or schema cannot support the check).
    def r001(self, row): return bool(UTI_RE.match(row["UTI"] or ""))
    def r002(self, row): return bool(UPI_RE.match(row["UPI"] or ""))
    def r003(self, row): return valid_lei_format(row["Counterparty1LEI"])
    def r004(self, row): return valid_lei_format(row["Counterparty2LEI"])

    def r005(self, row):
        n = is_number(row["NotionalAmount"])
        if n is None or n <= 0:
            return False
        whole, _, frac = row["NotionalAmount"].partition(".")
        return (len(whole.lstrip("-")) + len(frac)) <= 25 and len(frac) <= 5

    def r006(self, row):
        ccy = row["Currency"] or ""
        return bool(CCY_RE.match(ccy)) and ccy in ACTIVE_CCY

    def r007(self, row): return parse_date(row["MaturityDate"]) is not None
    def r008(self, row): return parse_dt(row["ReportingTimestamp"]) is not None
    def r009(self, row): return row["ActionType"] in ACTIONTYPE_ENUM
    def r010(self, row): return row["AssetClass"] in ASSETCLASS_ENUM
    def r011(self, row): return bool(row["UTI"])

    def r012(self, row):
        if row["ActionType"] not in UPI_MANDATORY_ACTIONS:
            return None
        return bool(row["UPI"])

    def r013(self, row): return bool(row["Counterparty1LEI"])
    def r014(self, row): return bool(row["Counterparty2LEI"])
    def r015(self, row): return bool(row["NotionalAmount"])
    def r016(self, row): return bool(row["Currency"])
    def r017(self, row): return bool(row["MaturityDate"])
    def r018(self, row): return bool(row["ReportingTimestamp"])
    def r019(self, row): return bool(row["ActionType"])
    def r020(self, row): return bool(row["AssetClass"])

    def r021(self, row):
        group = self.by_lifecycle.get(lifecycle_key(row["TradeID"]), [row])
        return len({r["UTI"] for r in group}) <= 1

    def r022(self, row):
        rec = self.dsb.get(row["UPI"])
        return rec is not None and rec.get("status") == "ACTIVE"

    def r023(self, row):
        rec = self.gleif.get(row["Counterparty1LEI"])
        return rec is not None and rec.get("registrationStatus") in ("ISSUED", "LAPSED")

    def r024(self, row):
        rec = self.gleif.get(row["Counterparty2LEI"])
        return rec is not None and rec.get("registrationStatus") in ("ISSUED", "LAPSED")

    def r025(self, row):
        if not row["NotionalAmount"]:
            return None  # covered by R015
        return bool(row["Currency"])

    def r026(self, row): return (row["Currency"] or "") in ACTIVE_CCY

    def r027(self, row):
        m, e = parse_date(row["MaturityDate"]), parse_dt(row["ExecutionTimestamp"])
        return None if not m or not e else m >= e.date()

    def r028(self, row):
        r, e = parse_dt(row["ReportingTimestamp"]), parse_dt(row["ExecutionTimestamp"])
        return None if not r or not e else r >= e

    def r029(self, row):
        if row["ActionType"] not in PRIOR_NEWT_REQUIRED_ACTIONS:
            return None
        return any(p["ActionType"] == "NEWT" for p in self.by_uti.get(row["UTI"], []))

    def r030(self, row):
        rec = self.dsb.get(row["UPI"])
        return None if rec is None else rec.get("assetClass") == row["AssetClass"]

    def r031(self, row):
        group = self.newt_by_generator.get((row["Counterparty1LEI"], row["UTI"]), [])
        if row["ActionType"] != "NEWT" or len(group) <= 1:
            return True
        return group[0] is row

    def r032(self, row):
        rec = self.dsb.get(row["UPI"])
        return None if rec is None else rec.get("assetClass") == row["AssetClass"]

    def r033(self, row): return row["Counterparty1LEI"] == SUBMITTING_ENTITY_LEI
    def r034(self, row): return row["Counterparty1LEI"] != row["Counterparty2LEI"]

    def r035(self, row):
        n = is_number(row["NotionalAmount"])
        return n is not None and n > 0

    def r036(self, row):
        ccy = row["Currency"]
        return bool(ccy) and ccy == ccy.upper()

    def r037(self, row):
        m, e = parse_date(row["MaturityDate"]), parse_dt(row["ExecutionTimestamp"])
        if not m or not e:
            return None
        try:
            limit = e.date().replace(year=e.year + 100)
        except ValueError:
            limit = e.date().replace(year=e.year + 100, day=28)
        return m <= limit

    def r038(self, row):
        r, e = parse_dt(row["ReportingTimestamp"]), parse_dt(row["ExecutionTimestamp"])
        return None if not r or not e else r <= e + timedelta(days=1)

    def r039(self, row):
        return None  # no LinkedUTIs field in schema -- not evaluable, see README

    def r040(self, row):
        if row["AssetClass"] != "CU":
            return None
        return bool(row["Currency"])

    def r041(self, row):
        uti, lei = row["UTI"] or "", row["Counterparty1LEI"] or ""
        return uti[:20] == lei

    def r042(self, row):
        rec = self.dsb.get(row["UPI"])
        rt = parse_dt(row["ReportingTimestamp"])
        if rec is None or rt is None:
            return None
        refresh = parse_date(rec.get("lastRefreshDate", ""))
        return None if refresh is None else (rt.date() - refresh).days <= 180

    def r043(self, row):
        rec = self.gleif.get(row["Counterparty1LEI"])
        rt = parse_dt(row["ReportingTimestamp"])
        if rec is None or rt is None or not rec.get("nextRenewalDate"):
            return True
        renewal = parse_date(rec["nextRenewalDate"])
        return True if renewal is None else not (0 <= (renewal - rt.date()).days <= 30)

    def r044(self, row):
        rec = self.gleif.get(row["Counterparty2LEI"])
        return True if rec is None else rec.get("registrationStatus") != "LAPSED"

    def r045(self, row):
        n = is_number(row["NotionalAmount"])
        return None if n is None else n <= 10_000_000_000

    def r046(self, row): return (row["Currency"] or "") not in DEPRECATED_CCY

    def r047(self, row):
        m = parse_date(row["MaturityDate"])
        return None if m is None else m.weekday() < 5

    def r048(self, row):
        r, e = parse_dt(row["ReportingTimestamp"]), parse_dt(row["ExecutionTimestamp"])
        if not r or not e:
            return None
        deadline = datetime.combine((e + timedelta(days=1)).date(), datetime.min.time()) + timedelta(hours=17)
        return not (deadline - timedelta(hours=2) <= r <= deadline)

    def r049(self, row):
        if row["ActionType"] not in ("CORR", "EROR"):
            return None
        peers = [p for p in self.by_uti.get(row["UTI"], []) if p["ActionType"] in ("CORR", "EROR")]
        return len(peers) <= 3

    def r050(self, row): return row["AssetClass"] != "OT"

    def run(self, rule_meta):
        results = []
        for row in self.trades:
            for rule_id, fn in self.rules.items():
                outcome = fn(row)
                sev = rule_meta[rule_id]["severity"]
                status = "N/A" if outcome is None else ("PASS" if outcome else ("FAIL" if sev == "ERROR" else "WARNING"))
                results.append({
                    "TradeID": row["TradeID"],
                    "RuleId": rule_id,
                    "Field": rule_meta[rule_id]["field"],
                    "Type": rule_meta[rule_id]["type"],
                    "Severity": sev,
                    "Status": status,
                    "Message": rule_meta[rule_id]["message"],
                })
        return results


# ---- Excel report ---------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
STATUS_FILL = {
    "FAIL": PatternFill("solid", fgColor="F8CBAD"),
    "WARNING": PatternFill("solid", fgColor="FFE699"),
    "PASS": PatternFill("solid", fgColor="C6E0B4"),
    "N/A": PatternFill("solid", fgColor="EDEDED"),
}


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    # NB: set freeze_panes with a string coordinate, not ws.cell(...) -- calling
    # ws.cell() to build that reference materializes a blank row and throws off
    # every subsequent ws.append() by one.
    ws.freeze_panes = f"A{row + 1}"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(results, trade_order):
    wb = Workbook()
    by_trade = defaultdict(list)
    for r in results:
        by_trade[r["TradeID"]].append(r)

    # Sheet 1: Trade Summary
    ws1 = wb.active
    ws1.title = "Trade Summary"
    ws1.append(["TradeID", "Overall Status", "Errors (FAIL)", "Warnings", "Failed Rule IDs", "Warning Rule IDs"])
    style_header(ws1, 6)
    for tid in trade_order:
        rows = by_trade[tid]
        fails = sorted(r["RuleId"] for r in rows if r["Status"] == "FAIL")
        warns = sorted(r["RuleId"] for r in rows if r["Status"] == "WARNING")
        overall = "FAIL" if fails else ("WARNING" if warns else "PASS")
        ws1.append([tid, overall, len(fails), len(warns), ", ".join(fails), ", ".join(warns)])
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        row[1].fill = STATUS_FILL.get(row[1].value, STATUS_FILL["N/A"])
    autosize(ws1, [34, 15, 13, 11, 40, 40])
    ws1.auto_filter.ref = f"A1:F{ws1.max_row}"

    # Sheet 2: Failures And Warnings
    ws2 = wb.create_sheet("Failures And Warnings")
    ws2.append(["TradeID", "RuleId", "Field", "Type", "Severity", "Status", "Message"])
    style_header(ws2, 7)
    for r in results:
        if r["Status"] in ("FAIL", "WARNING"):
            ws2.append([r["TradeID"], r["RuleId"], r["Field"], r["Type"], r["Severity"], r["Status"], r["Message"]])
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        row[5].fill = STATUS_FILL.get(row[5].value, STATUS_FILL["N/A"])
        row[6].alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws2, [34, 9, 20, 13, 10, 10, 90])
    ws2.auto_filter.ref = f"A1:G{ws2.max_row}"

    # Sheet 3: Rule Summary
    ws3 = wb.create_sheet("Rule Summary")
    ws3.append(["RuleId", "Field", "Type", "Severity", "PASS", "FAIL", "WARNING", "N/A", "Message"])
    style_header(ws3, 9)
    by_rule = defaultdict(lambda: {"PASS": 0, "FAIL": 0, "WARNING": 0, "N/A": 0})
    rule_meta_row = {}
    for r in results:
        by_rule[r["RuleId"]][r["Status"]] += 1
        rule_meta_row[r["RuleId"]] = (r["Field"], r["Type"], r["Severity"], r["Message"])
    for rid in sorted(by_rule):
        field, typ, sev, msg = rule_meta_row[rid]
        c = by_rule[rid]
        ws3.append([rid, field, typ, sev, c["PASS"], c["FAIL"], c["WARNING"], c["N/A"], msg])
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        row[8].alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws3, [9, 20, 13, 10, 8, 8, 10, 8, 90])
    ws3.auto_filter.ref = f"A1:I{ws3.max_row}"

    # Sheet 4: All Results
    ws4 = wb.create_sheet("All Results")
    ws4.append(["TradeID", "RuleId", "Field", "Type", "Severity", "Status", "Message"])
    style_header(ws4, 7)
    for r in results:
        ws4.append([r["TradeID"], r["RuleId"], r["Field"], r["Type"], r["Severity"], r["Status"], r["Message"]])
    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
        row[5].fill = STATUS_FILL.get(row[5].value, STATUS_FILL["N/A"])
        row[6].alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws4, [34, 9, 20, 13, 10, 10, 90])
    ws4.auto_filter.ref = f"A1:G{ws4.max_row}"

    return wb


# ---- main ---------------------------------------------------------------------
def find_project_root() -> Path:
    """Locate the project root by searching upward from cwd for reference_data/.

    Not derived from this script's own file location: the skill may be
    invoked from a copy living outside the project (e.g. the user-level
    skills folder), in which case a path derived from __file__ would point
    at the wrong place. Falls back to cwd if no marker is found so relative
    --trades/--rules/etc. flags still resolve predictably.
    """
    cur = Path.cwd().resolve()
    for candidate in (cur, *cur.parents):
        if (candidate / "reference_data" / "validation_rules.json").exists():
            return candidate
    return cur


def resolve_trades_arg(value: str, project_root: Path) -> str:
    """Resolve --trades so a bare filename works from any cwd.

    An absolute path, or a path that exists as given (relative to cwd), is
    used as-is. A bare filename that doesn't exist relative to cwd falls
    back to <project_root>/<value>, so "--trades foo.csv" (or a filename
    passed as the skill's parameter) finds the file next to
    trade_data_200.csv even if the shell's cwd differs from the project root.
    """
    p = Path(value)
    if p.is_absolute() or p.exists():
        return value
    candidate = project_root / p
    return str(candidate) if candidate.exists() else value


def main():
    project_root = find_project_root()

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--trades", default=str(project_root / "trade_data_200.csv"),
                     help="Trade CSV to validate. Bare filenames are also looked up "
                          "under the project root if not found relative to cwd.")
    ap.add_argument("--rules", default=str(project_root / "reference_data" / "validation_rules.json"))
    ap.add_argument("--gleif", default=str(project_root / "reference_data" / "gleif_lei_reference.json"))
    ap.add_argument("--dsb", default=str(project_root / "reference_data" / "dsb_upi_reference.json"))
    ap.add_argument("--out", default=str(project_root / "test_results.xlsx"))
    args = ap.parse_args()
    args.trades = resolve_trades_arg(args.trades, project_root)

    with open(args.rules, encoding="utf-8") as f:
        rules = json.load(f)
    rule_meta = {r["ruleId"]: r for r in rules}

    with open(args.trades, newline="", encoding="utf-8") as f:
        trades = list(csv.DictReader(f))
    if not trades:
        sys.exit(f"No trade rows found in {args.trades}")

    with open(args.gleif, encoding="utf-8") as f:
        gleif = {r["lei"]: r for r in json.load(f)["records"]}

    with open(args.dsb, encoding="utf-8") as f:
        dsb = {r["upi"]: r for r in json.load(f)["records"]}

    engine = Engine(trades, gleif, dsb)
    results = engine.run(rule_meta)

    trade_order, seen = [], set()
    for row in trades:
        if row["TradeID"] not in seen:
            seen.add(row["TradeID"])
            trade_order.append(row["TradeID"])

    wb = build_workbook(results, trade_order)
    wb.save(args.out)

    n_fail = sum(1 for r in results if r["Status"] == "FAIL")
    n_warn = sum(1 for r in results if r["Status"] == "WARNING")
    n_pass = sum(1 for r in results if r["Status"] == "PASS")
    n_na = sum(1 for r in results if r["Status"] == "N/A")
    trades_fail = sum(1 for tid in trade_order if any(r["TradeID"] == tid and r["Status"] == "FAIL" for r in results))

    print(f"Evaluated {len(trades)} trades x {len(rule_meta)} rules = {len(results)} checks")
    print(f"PASS={n_pass}  FAIL={n_fail}  WARNING={n_warn}  N/A={n_na}")
    print(f"Trades with at least one ERROR: {trades_fail} / {len(trade_order)}")
    print(f"Workbook written to: {args.out}")


if __name__ == "__main__":
    main()
